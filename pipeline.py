"""
pipeline.py — Pipeline consolidada: SIGLA → dadosManut → Email

Fluxo:
  0. Leitura dos trilhos via SIGLA (sigla_automacao_v2.py) — 1 data por vez
  1. Resolve datas alvo (mesma lógica do BIFrotaManut main.py)
  2. Para cada data: processa relatórios SIGLA_Relatorio_*.xlsx
     e insere registros novos diretamente em dadosManut.xlsx
  3. Executa rotina BIFrotaManut (gera PNGs e envia email via Outlook)

Pré-requisito:
  - SIGLA aberto e visível na tela (para a etapa 0)
  - dadosManut.xlsx acessível em BIFrotaManut - 2/dadosManut.xlsx
  - Outlook instalado e configurado (para envio de email)

Uso:
  python pipeline.py                              # D-1 (ou sex+sáb+dom na segunda)
  python pipeline.py --data 2026-06-01            # 1 dia específico
  python pipeline.py --data 2026-05-30 --ate 2026-06-01  # período
  python pipeline.py --from-malha                 # SIGLA já aberto na malha certa
  python pipeline.py --pular-sigla                # pula etapa 0 (extração já feita)
  python pipeline.py --dry-run                    # gera PNGs, sem email
  python pipeline.py --enviar                     # força envio imediato
"""
from __future__ import annotations

import argparse
import logging
import os
import subprocess
import sys
from datetime import date, datetime, time, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---------- Caminhos raiz ----------
ROOT          = Path(__file__).parent
SIGLA_DIR     = ROOT / "ControleRemoto SIGLA"
BIFFROTA_DIR  = ROOT / "BIFrotaManut - 2"
DADOS_MANUT   = BIFFROTA_DIR / "dadosManut.xlsx"
CONFIG_YAML   = BIFFROTA_DIR / "automacao" / "config.yaml"

# Coloca os dois módulos no path
sys.path.insert(0, str(SIGLA_DIR))
sys.path.insert(0, str(BIFFROTA_DIR / "automacao"))


# =========================================================
# Logging
# =========================================================

def _setup_logging(log_dir: Path) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    arq = log_dir / f"pipeline_{datetime.now():%Y%m%d_%H%M%S}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[
            logging.FileHandler(arq, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


# =========================================================
# Resolução de datas (mesma lógica do main.py BIFrotaManut)
# =========================================================

def _parse_data(s: str) -> date:
    s = s.strip().replace("/", "-")
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Data inválida: '{s}'. Use YYYY-MM-DD ou DD/MM/YYYY")


def resolver_datas(override=None, ate_override=None) -> list[date]:
    """Retorna lista de datas alvo com base nos argumentos (ou D-1 por padrão)."""
    if override and ate_override:
        d1, d2 = _parse_data(override), _parse_data(ate_override)
        if d2 < d1:
            d1, d2 = d2, d1
        return [d1 + timedelta(days=i) for i in range((d2 - d1).days + 1)]
    if override:
        return [_parse_data(override)]
    # Padrão: D-1; na segunda-feira cobre sex+sáb+dom
    hoje = date.today()
    if hoje.weekday() == 0:  # segunda
        sexta = hoje - timedelta(days=3)
        return [sexta, sexta + timedelta(1), sexta + timedelta(2)]
    return [hoje - timedelta(days=1)]


# =========================================================
# Etapa 0 — Leitura dos trilhos via sigla_automacao_v2.py
# =========================================================

def etapa_leitura_sigla(datas: list[date], from_malha: bool, log) -> bool:
    """
    Chama sigla_automacao_v2.py via subprocess para cada data.
    - from_malha=True: assume SIGLA já aberto na malha certa (pula login/filtro)
    - Retorna True se todas as datas foram processadas com sucesso.
    """
    script = SIGLA_DIR / "sigla_automacao_v2.py"
    config = SIGLA_DIR / "config.json"

    if not script.exists():
        raise FileNotFoundError(f"Script não encontrado: {script}")
    if not config.exists():
        raise FileNotFoundError(
            f"config.json não encontrado em {SIGLA_DIR}. "
            "Execute a calibração (EXECUTAR.bat → Opção 2) antes de usar a pipeline."
        )

    sucesso = True
    primeira = True
    for d in datas:
        data_sigla = d.strftime("%d/%m/%Y")
        cmd = [sys.executable, str(script), "--config", str(config), "--data", data_sigla]

        if from_malha:
            # Usuário disse que a malha já está aberta na data certa
            cmd.append("--from-malha")
            modo = "from-malha"
        elif not primeira:
            # 2ª data em diante: SIGLA já aberto, só reapre o filtro Alt+C
            cmd.append("--sigla-aberto")
            modo = "sigla-aberto"
        else:
            modo = "completo"

        log.info("  Lendo trilhos SIGLA para %s (modo: %s)...", data_sigla, modo)
        log.info("  Comando: %s", " ".join(cmd))

        resultado = subprocess.run(cmd, cwd=str(SIGLA_DIR))
        if resultado.returncode != 0:
            log.error("  [ERRO] sigla_automacao_v2.py retornou código %d para %s",
                      resultado.returncode, data_sigla)
            sucesso = False
        else:
            log.info("  [OK] Leitura concluída para %s.", data_sigla)

        primeira = False

    return sucesso


# =========================================================
# Etapa 1 — Processar relatórios SIGLA e inserir em dadosManut
# =========================================================

def inserir_em_dadosmanut(novos_registros: list[dict], log) -> int:
    """
    Insere registros novos diretamente na aba DADOS do dadosManut.xlsx.
    Retorna quantidade de linhas inseridas.
    """
    if not novos_registros:
        return 0

    COLUNAS = ["Data", "Dia Sem", "Local", "Carro", "Frota", "Empresa",
               "H.Ini.", "H.Fim", "Duracao", "Tipo"]

    wb = load_workbook(DADOS_MANUT)
    if "DADOS" not in wb.sheetnames:
        raise ValueError("Aba 'DADOS' não encontrada em dadosManut.xlsx")
    ws = wb["DADOS"]

    # Descobre a ordem real das colunas pelo cabeçalho
    header = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else ""
              for c in range(1, ws.max_column + 1)]
    col_idx = {nome: i + 1 for i, nome in enumerate(header)}

    # Mapeia nomes com/sem acento
    aliases = {"Duracao": ["Duracao", "Duração"]}
    for base, opts in aliases.items():
        for opt in opts:
            if opt in col_idx and base not in col_idx:
                col_idx[base] = col_idx[opt]

    primeira_linha_vazia = ws.max_row + 1

    for i, rec in enumerate(novos_registros):
        row = primeira_linha_vazia + i

        def _set(col_nome, valor, fmt=None):
            if col_nome not in col_idx:
                return
            cell = ws.cell(row=row, column=col_idx[col_nome], value=valor)
            if fmt:
                cell.number_format = fmt

        data_val = datetime.combine(rec["Data"], time(0, 0))
        _set("Data",     data_val,       "DD/MM/YYYY")
        _set("Dia Sem",  rec["Dia Sem"])
        _set("Local",    rec["Local"])
        _set("Carro",    rec["Carro"])
        _set("Frota",    rec["Frota"])
        _set("Empresa",  rec["Empresa"])
        _set("H.Ini.",   rec["H.Ini."],  "HH:MM:SS")
        _set("H.Fim",    rec["H.Fim"],   "HH:MM:SS")
        _set("Duracao",  rec["Duracao"], "HH:MM:SS")
        _set("Tipo",     rec["Tipo"])

    wb.save(DADOS_MANUT)
    log.info("dadosManut.xlsx atualizado: %d linhas inseridas.", len(novos_registros))
    return len(novos_registros)


def etapa_exportar_manut(datas: list[date], log) -> int:
    """
    Importa e chama funções de exportar_manut.py para processar as datas
    e insere os resultados diretamente em dadosManut.xlsx.
    Retorna total de registros inseridos.
    """
    # Importa módulo do Processo 1
    try:
        import exportar_manut as em
    except ImportError as e:
        raise RuntimeError(
            f"Não foi possível importar exportar_manut.py de {SIGLA_DIR}: {e}"
        ) from e

    # Substitui o caminho DADOS_MANUT do módulo pelo correto
    em.DADOS_MANUT = DADOS_MANUT
    em.SAIDA_DIR   = SIGLA_DIR / "saida"

    # Datas já presentes (para não duplicar)
    ja_existentes = em.datas_existentes_no_manut()
    log.info("Datas já existentes em dadosManut: %d", len(ja_existentes))

    todos = []
    for d in datas:
        if d in ja_existentes:
            log.info("  [PULADO] %s já existe em dadosManut.", d.isoformat())
            continue
        log.info("  Processando %s...", d.isoformat())
        recs = em.processar_dia(d)
        log.info("    → %d registro(s)", len(recs))
        todos.extend(recs)

    if not todos:
        log.info("Nenhum registro novo para inserir.")
        return 0

    return inserir_em_dadosmanut(todos, log)


# =========================================================
# Etapa 1.5 — Atualizar dadosManut com PDFs de manutenção
# =========================================================

def etapa_atualizar_manut(datas: list[date], log) -> None:
    """
    Lê os PDFs de manutenção da pasta Downloads e reconcilia com dadosManut.xlsx.
    Aplica as regras: Reten.→VTR (amarelo), VTR→Reten. (verde), insere novos.

    Se não encontrar PDFs correspondentes às datas alvo, pausa e pergunta
    se deve continuar mesmo assim.
    """
    try:
        import atualizar_manut as am
    except ImportError as e:
        raise RuntimeError(
            f"Não foi possível importar atualizar_manut.py de {SIGLA_DIR}: {e}"
        ) from e

    pasta_pdfs = am.DEFAULT_PDF_FOLDER
    excel      = str(DADOS_MANUT)

    # Verifica se há PDFs cujas datas (extraídas do nome) batem com as datas alvo
    todos_pdfs = [f for f in os.listdir(pasta_pdfs) if f.lower().endswith('.pdf')]
    datas_str  = {d.strftime("%d.%m") for d in datas}  # ex: {"01.06"}
    pdfs_da_data = [f for f in todos_pdfs
                    if any(ds in f for ds in datas_str)]

    if not pdfs_da_data:
        log.warning("Nenhum PDF encontrado em '%s' para as datas: %s",
                    pasta_pdfs, [d.isoformat() for d in datas])
        print()
        print("=" * 60)
        print("  ATENÇÃO: Nenhum PDF de manutenção encontrado para")
        print(f"  {', '.join(d.strftime('%d/%m/%Y') for d in datas)}")
        print(f"  Pasta verificada: {pasta_pdfs}")
        print()
        resp = input("  Deseja continuar sem os PDFs? [s/N]: ").strip().lower()
        print("=" * 60)
        print()
        if resp not in ("s", "sim", "y", "yes"):
            raise RuntimeError(
                "Pipeline interrompida pelo usuário: PDFs de manutenção não encontrados."
            )
        log.info("Usuário optou por continuar sem PDFs.")
        return

    log.info("PDFs encontrados para as datas alvo: %s", pdfs_da_data)
    am.processar(pasta_pdfs, excel)
    log.info("Reconciliação com PDFs de manutenção concluída.")


# =========================================================
# Etapa 2 — BIFrotaManut: gerar PNGs e enviar email
# =========================================================

def etapa_bifrotamanut(datas: list[date], forcar_envio: bool, dry_run: bool, log) -> int:
    """
    Chama main.executar() do BIFrotaManut com as datas alvo.
    """
    try:
        import main as bifmain
    except ImportError as e:
        raise RuntimeError(
            f"Não foi possível importar main.py de {BIFFROTA_DIR / 'automacao'}: {e}"
        ) from e

    log.info("Iniciando BIFrotaManut (dry_run=%s, enviar=%s)...", dry_run, forcar_envio)

    # Sempre passa data_ate para bifmain, mesmo com 1 data.
    # Sem isso, bifmain._resolver_datas chama _datas_a_partir_de() que expande
    # domingo para sex+sáb+dom — ignorando a data exata que o usuário escolheu.
    # Com data_ini == data_ate, o caminho de período retorna exatamente [data_ini].
    data_ini = datas[0].isoformat()
    data_ate = datas[-1].isoformat()

    codigo = bifmain.executar(
        config_path=CONFIG_YAML,
        data_override=data_ini,
        data_ate_override=data_ate,
        forcar_envio=forcar_envio,
        dry_run=dry_run,
    )
    return codigo


# =========================================================
# Orquestrador principal
# =========================================================

def executar(data_override=None, data_ate_override=None,
             forcar_envio=False, dry_run=False,
             pular_sigla=False, from_malha=False) -> int:

    log_dir = BIFFROTA_DIR / "automacao" / "logs"
    _setup_logging(log_dir)
    log = logging.getLogger("pipeline")

    datas = resolver_datas(data_override, data_ate_override)
    log.info("=== Pipeline consolidada iniciada ===")
    log.info("Datas alvo: %s", [d.isoformat() for d in datas])

    # --- Etapa 0: leitura dos trilhos via SIGLA ---
    if pular_sigla:
        log.info("--- Etapa 0: PULADA (--pular-sigla) ---")
    else:
        log.info("--- Etapa 0: Leitura dos trilhos SIGLA ---")
        log.info("NAO MEXA NO MOUSE durante a varredura.")
        try:
            ok = etapa_leitura_sigla(datas, from_malha, log)
            if not ok:
                log.error("Etapa 0 terminou com erros em uma ou mais datas. Abortando.")
                return 1
            log.info("Etapa 0 concluída.")
        except Exception:
            log.exception("Falha na Etapa 0 (leitura SIGLA). Abortando.")
            return 1

    # --- Etapa 1: exportar SIGLA → dadosManut ---
    log.info("--- Etapa 1: SIGLA → dadosManut ---")
    try:
        inseridos = etapa_exportar_manut(datas, log)
        log.info("Etapa 1 concluída. %d registro(s) inserido(s).", inseridos)
    except Exception:
        log.exception("Falha na Etapa 1 (exportar_manut). Abortando.")
        return 1

    # --- Etapa 1.5: reconciliar com PDFs de manutenção ---
    log.info("--- Etapa 1.5: atualizar dadosManut com PDFs de manutenção ---")
    try:
        etapa_atualizar_manut(datas, log)
        log.info("Etapa 1.5 concluída.")
    except Exception:
        log.exception("Falha na Etapa 1.5 (atualizar_manut). Abortando.")
        return 1

    # --- Etapa 2: BIFrotaManut ---
    log.info("--- Etapa 2: BIFrotaManut (gerar PNGs + email) ---")
    try:
        codigo = etapa_bifrotamanut(datas, forcar_envio, dry_run, log)
    except Exception:
        log.exception("Falha na Etapa 2 (BIFrotaManut).")
        return 2

    log.info("=== Pipeline concluída (código=%d) ===", codigo)
    return codigo


def _parse_args():
    ap = argparse.ArgumentParser(description="Pipeline consolidada SIGLA → dadosManut → Email")
    ap.add_argument("--data", default=None,
                    help="Data alvo YYYY-MM-DD (padrão: D-1)")
    ap.add_argument("--ate", default=None,
                    help="Data final para período (usar com --data)")
    ap.add_argument("--from-malha", action="store_true",
                    help="SIGLA já aberto na malha certa (pula login e filtro de data)")
    ap.add_argument("--pular-sigla", action="store_true",
                    help="Pula etapa 0 (extração SIGLA já foi feita)")
    ap.add_argument("--enviar", action="store_true",
                    help="Força envio imediato do email (sobrepõe config.yaml)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Apenas insere em dadosManut e gera PNGs, sem email")
    return ap.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    sys.exit(executar(
        data_override=args.data,
        data_ate_override=args.ate,
        forcar_envio=args.enviar,
        dry_run=args.dry_run,
        pular_sigla=args.pular_sigla,
        from_malha=args.from_malha,
    ))
